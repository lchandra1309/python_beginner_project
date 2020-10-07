import openpyxl 

path=r".//openpyxl_excersise/openpyxl.xlsx"
wb=openpyxl.load_workbook(path)

sheet=wb['lchandra']
print(sheet.title)

# data=[
#     ('lokeshkumar',25,'annotator',1.2,'python'),
#     ('malini',35,'validator',2.5,'java'),
#     ('sudhakar',25,'developer',2,'java'),   
#     ('pragadeesh',25,'testengineer',5,'c#')
# ]
# for i in data:
#     sheet.append(i)

# for i in range(4,8):
#     sheet.delete_rows(i)

# data=('pragadeesh',25,'test engineer',2,'ruby')
# for i,j in enumerate(data):
#     sheet.cell(row=6,column=i+1,value=j)

# for i in range(1,sheet.max_row+1):
#     for j in range(1,sheet.max_column+1):
#         print(sheet.cell(i,j).value)
#     print('*'*10)

# chg=sheet.cell(row=2,column=3)
# chg.value='Team Lead'


num=sheet['a:b']
print(num)
wb.save(path)